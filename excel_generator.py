"""
Excel Invoice generator: uses the template Excel file and fills in
dynamic content (invoice number, date, buyer, products, total).
All static content stays exactly as in the template.
"""
from __future__ import annotations

import os
import re
import time
import json
import random
import asyncio
from datetime import datetime
from typing import Any
import pytz

import aiohttp
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.utils import range_boundaries
from copy import copy

from config import OUTPUT_DIR, PRODUCTS_JSON_URL, TEMPLATE_PATH

INVOICE_NUMBERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "invoice_numbers.json")

# Cache for the products payload
_PRODUCTS_CACHE: dict[str, Any] | None = None
_PRODUCTS_CACHE_TS: float = 0.0
_PRODUCTS_LOCK = asyncio.Lock()
CACHE_TTL = 600  # 10 minutes

# Template constants (original template rows)
SHEET_NAME = "ინვოისი"
HEADER_ROW = 12
# After inserting a spacer row at 13, all product/bottom rows shift +1
SPACER_ROW = 13
FIRST_DATA_ROW = 14
LAST_DATA_ROW = 24
TEMPLATE_ROWS = LAST_DATA_ROW - FIRST_DATA_ROW + 1  # 11 rows in template
MAX_ITEMS = 33  # max products per invoice (fits one page)
BANK_DETAILS_ROW = 26
GRAND_TOTAL_ROW = 26
DIRECTOR_ROW = 38
SIGNATURE_ROW = 40
FOOTER_ROW = 43

# Colors from template
HEADER_FILL = "00818A"
ROW_ALT_FILL = "F7FAFC"
GRAND_TOTAL_FILL = "EBF8FA"


async def _load_products(force: bool = False) -> Any:
    """Load products from the shared local cache file.

    The cache is refreshed by a cron job every 30 minutes (see
    scraper_common/products_cache.py).  The invoice bot never hits
    acoustic.ge directly – it always reads from the local cache.
    """
    global _PRODUCTS_CACHE, _PRODUCTS_CACHE_TS
    async with _PRODUCTS_LOCK:
        expired = (time.monotonic() - _PRODUCTS_CACHE_TS) > CACHE_TTL
        if _PRODUCTS_CACHE is None or expired or force:
            try:
                from products_cache import load_products, cache_age_seconds
            except ImportError:
                raise RuntimeError(
                    "scraper_common not on PYTHONPATH – cannot load products cache. "
                    "Set PYTHONPATH=/root/scraper_common"
                )

            age = cache_age_seconds()
            data = load_products(allow_fetch=False)
            if data is None:
                raise RuntimeError(
                    "Products cache is empty. Run: python /root/scraper_common/products_cache.py --refresh"
                )
            _PRODUCTS_CACHE = data
            _PRODUCTS_CACHE_TS = time.monotonic()
            if age is not None:
                print(f"[invoice] loaded products from cache (age: {age:.0f}s)", flush=True)
        return _PRODUCTS_CACHE


def _iter_products(payload: Any):
    def _is_active(item: dict) -> bool:
        return item.get("status", "A") != "D"

    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict) and _is_active(item):
                yield item
    elif isinstance(payload, dict):
        if "products" in payload and isinstance(payload["products"], list):
            for item in payload["products"]:
                if isinstance(item, dict) and _is_active(item):
                    yield item
            return
        for key, item in payload.items():
            if isinstance(item, dict) and _is_active(item):
                item.setdefault("_key", key)
                yield item


_PRODUCT_ID_RE = re.compile(r"product_id=(\d+)")


def _extract_ids(product: dict) -> list[str]:
    ids: list[str] = []
    sku = product.get("sku")
    if sku not in (None, ""):
        ids.append(str(sku).strip())

    url = product.get("url") or product.get("URL") or ""
    if isinstance(url, str):
        m = _PRODUCT_ID_RE.search(url)
        if m:
            ids.append(m.group(1))

    for k in ("product_id", "productId", "id", "ID", "Id", "code", "_key"):
        v = product.get(k)
        if v not in (None, ""):
            ids.append(str(v).strip())

    seen: set[str] = set()
    out: list[str] = []
    for i in ids:
        if i not in seen:
            seen.add(i)
            out.append(i)
    return out


def _extract_name(product: dict) -> str:
    for k in ("product", "title", "name", "full_name", "product_name", "Name", "Title"):
        v = product.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _extract_price(product: dict) -> float:
    for k in ("price", "Price", "unit_price", "sale_price", "current_price", "final_price"):
        v = product.get(k)
        if v in (None, ""):
            continue
        try:
            if isinstance(v, str):
                cleaned = re.sub(r"[^\d.,-]", "", v).replace(",", ".")
                return float(cleaned)
            return float(v)
        except (TypeError, ValueError):
            continue
    return 0.0


def _normalize_code(code: str) -> str:
    return re.sub(r'[^A-Za-z0-9]', '', code).upper()


async def fetch_product_data(product_id: str | int) -> dict:
    target = str(product_id).strip()
    if not target:
        raise ValueError("Product code cannot be empty")

    target_norm = _normalize_code(target)
    payload = await _load_products()

    products_with_ids = []
    for product in _iter_products(payload):
        for pid in _extract_ids(product):
            pid_norm = _normalize_code(pid)
            products_with_ids.append({
                "id": pid,
                "id_norm": pid_norm,
                "name": _extract_name(product),
                "price": _extract_price(product),
            })

    for item in products_with_ids:
        if item["id_norm"] == target_norm:
            return {
                "id": item["id"],
                "name": item["name"],
                "price": item["price"],
            }

    candidates = []
    for item in products_with_ids:
        if item["id_norm"].startswith(target_norm):
            candidates.append(item)

    if not candidates:
        raise ValueError(f"Error: Product code {target} not found in database.")

    if len(candidates) > 1:
        variant_list = ", ".join([f"{c['id']} ({c['name']})" for c in candidates])
        raise ValueError(f"Found multiple variants for this code: {variant_list}. Please specify which one you meant.")

    return {
        "id": candidates[0]["id"],
        "name": candidates[0]["name"],
        "price": candidates[0]["price"],
    }


def _filename_client_name(client_info: str) -> str:
    name = re.sub(r"\s+\(\d+/\d+\)\s*$", "", client_info.strip())
    return re.sub(r"\s+\d{9}\s*$", "", name).strip() or "Client"


def _sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\r\n\t]+", " ", name).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "Client"


def _generate_invoice_number() -> str:
    used = set()
    if os.path.exists(INVOICE_NUMBERS_FILE):
        try:
            with open(INVOICE_NUMBERS_FILE, "r", encoding="utf-8") as f:
                used = set(json.load(f))
        except (json.JSONDecodeError, IOError):
            pass

    for _ in range(10000):
        num = f"INV-{random.randint(100000, 999999)}"
        if num not in used:
            used.add(num)
            with open(INVOICE_NUMBERS_FILE, "w", encoding="utf-8") as f:
                json.dump(sorted(used), f, indent=2)
            return num

    raise RuntimeError("Could not generate a unique invoice number after 10000 attempts.")


def _fmt_price(val: float) -> float:
    return round(val, 2)


def _insert_rows_safe(ws, at: int, count: int) -> None:
    """Insert `count` rows at position `at`, shifting merged cells, row heights,
    and image anchors that openpyxl doesn't handle automatically."""
    # Save merged ranges at or below the insertion point, unmerge them
    shifted_merges = []
    for rng in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        if min_row >= at:
            ws.unmerge_cells(str(rng))
            shifted_merges.append((min_col, min_row, max_col, max_row))

    ws.insert_rows(at, count)

    # Re-merge shifted ranges
    for min_col, min_row, max_col, max_row in shifted_merges:
        ws.merge_cells(start_row=min_row + count, start_column=min_col,
                       end_row=max_row + count, end_column=max_col)

    # Shift row heights below insertion point
    heights = {r: dim.height for r, dim in ws.row_dimensions.items()
               if dim.height is not None and r >= at}
    for r in sorted(heights, reverse=True):
        ws.row_dimensions[r + count].height = heights[r]
        ws.row_dimensions[r].height = None

    # Shift image anchors below insertion point (0-indexed rows)
    for img in ws._images:
        anchor = img.anchor
        if hasattr(anchor, "_from") and anchor._from.row >= at - 1:
            anchor._from.row += count
            if hasattr(anchor, "to") and anchor.to is not None:
                anchor.to.row += count


def _insert_spacer(ws) -> None:
    """Insert a thin empty row between the header and product rows."""
    _insert_rows_safe(ws, SPACER_ROW, 1)
    ws.row_dimensions[SPACER_ROW].height = 21
    ws.row_dimensions[HEADER_ROW].height = 18


def _extend_table(ws, extra: int) -> None:
    """Insert `extra` product rows after LAST_DATA_ROW, copying template row styles."""
    insert_at = LAST_DATA_ROW + 1
    _insert_rows_safe(ws, insert_at, extra)

    # Style new rows by copying from template rows (alternating pattern)
    for i in range(extra):
        new_row = insert_at + i
        row_index_in_table = TEMPLATE_ROWS + i
        src_row = FIRST_DATA_ROW + (row_index_in_table % 2)
        for col in range(1, 7):
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=new_row, column=col)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
        ws.merge_cells(start_row=new_row, start_column=2, end_row=new_row, end_column=3)
        ws.row_dimensions[new_row].height = ws.row_dimensions[FIRST_DATA_ROW].height


async def generate_invoice(client_info: str, items: list[dict]) -> str:
    """
    Generate an Excel invoice by filling in the template.

    items: [{'id': '13333', 'qty': 2}, ...]
    Returns absolute path of the generated .xlsx file.
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template Excel file not found: {TEMPLATE_PATH}")

    # Fetch all products concurrently
    await _load_products()
    resolved = await asyncio.gather(
        *(fetch_product_data(it["id"]) for it in items),
        return_exceptions=True
    )

    for i, result in enumerate(resolved):
        if isinstance(result, Exception):
            raise result

    # Generate invoice number
    invoice_num = _generate_invoice_number()

    # Date
    tbilisi_tz = pytz.timezone('Asia/Tbilisi')
    current_date = datetime.now(tbilisi_tz).strftime("%d/%m/%Y")

    # Load template
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    # Insert spacer row between header and products
    _insert_spacer(ws)

    # Extend the table if more items than template rows
    num_items = len(items)
    extra = max(0, num_items - TEMPLATE_ROWS)
    if extra:
        _extend_table(ws, extra)
    last_data_row = LAST_DATA_ROW + extra

    # 1. Update invoice number + date (D2)
    ws.cell(row=2, column=4, value=f"ინვოისი  # {invoice_num}  |  თარიღი: {current_date}")

    # 2. Update buyer info (D6)
    ws.cell(row=6, column=4, value=client_info)

    # 3. Fill product rows
    for idx, (item, prod) in enumerate(zip(items, resolved)):
        row = FIRST_DATA_ROW + idx
        qty = item.get("qty", 0)
        try:
            qty = int(qty)
        except (TypeError, ValueError):
            qty = 0

        price = _fmt_price(prod["price"])

        # A column: row number (already set in template, but ensure)
        ws.cell(row=row, column=1, value=idx + 1)

        # B column (merged B:C): product name
        ws.cell(row=row, column=2, value=prod["name"])

        # D column: quantity
        ws.cell(row=row, column=4, value=qty)

        # E column: price
        ws.cell(row=row, column=5, value=price)

        # F column: line total (computed value, not formula — openpyxl doesn't calculate)
        line_total = _fmt_price(qty * price)
        ws.cell(row=row, column=6, value=line_total)

    # 4. Hide unused product rows so the lower sections move up dynamically
    for row in range(FIRST_DATA_ROW, last_data_row + 1):
        ws.row_dimensions[row].hidden = row >= FIRST_DATA_ROW + num_items

    # 5. Grand total (computed value — openpyxl doesn't calculate formulas)
    grand_total = sum(
        _fmt_price(int(it.get("qty", 0)) * _fmt_price(prod["price"]))
        for it, prod in zip(items, resolved)
    )
    ws.cell(row=GRAND_TOTAL_ROW + extra, column=6, value=grand_total)

    # 6. Make review.acoustic.ge a clickable hyperlink (shifted by spacer + extra rows)
    review_cell = ws.cell(row=32 + extra, column=1)
    if review_cell.value and "review.acoustic.ge" in str(review_cell.value):
        review_cell.hyperlink = "https://review.acoustic.ge"
        review_cell.font = Font(
            name=review_cell.font.name or "Quattrocento Sans",
            size=review_cell.font.size or 9,
            bold=False,
            color="FF0563C1",
            underline="single",
        )

    # Save
    safe_name = _sanitize_filename(_filename_client_name(client_info))
    filename_invoice_num = invoice_num.replace("-", "_")
    out_path = os.path.abspath(os.path.join(OUTPUT_DIR, f"{filename_invoice_num}_{safe_name}.xlsx"))
    wb.save(out_path)
    wb.close()
    return out_path
