"""
Invoice generator: fetches live product data from acoustic.ge JSON feed
and populates the master Excel template using a clean-write-save approach.
"""
from __future__ import annotations

import os
import re
import time
import json
import random
import asyncio
from datetime import datetime
from typing import Any
import pytz

import aiohttp
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import MergedCell

from config import TEMPLATE_PATH, OUTPUT_DIR, PRODUCTS_JSON_URL

INVOICE_NUMBERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "invoice_numbers.json")


def _safe_cell(ws, row, col):
    """Return the writable cell at (row, col), routing MergedCell to the range's top-left."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def _write_cell(ws, row, col, value=None, number_format=None):
    """Write value / number_format to a cell, transparently handling merged ranges."""
    cell = _safe_cell(ws, row, col)
    if value is not None:
        cell.value = value
    if number_format is not None:
        cell.number_format = number_format

# Cache for the products payload so repeated lookups inside one
# generate_invoice() call don't hammer the API.
# Expires after CACHE_TTL seconds so long-running bot always serves fresh prices.
_PRODUCTS_CACHE: dict[str, Any] | None = None
_PRODUCTS_CACHE_TS: float = 0.0
_PRODUCTS_LOCK = asyncio.Lock()
CACHE_TTL = 600  # 10 minutes


async def _load_products(force: bool = False) -> Any:
    """Fetch the full products JSON and cache it in memory (with TTL)."""
    global _PRODUCTS_CACHE, _PRODUCTS_CACHE_TS
    async with _PRODUCTS_LOCK:
        expired = (time.monotonic() - _PRODUCTS_CACHE_TS) > CACHE_TTL
        if _PRODUCTS_CACHE is None or expired or force:
            timeout = aiohttp.ClientTimeout(total=30)
            async with aiohttp.ClientSession(timeout=timeout) as session:
                async with session.get(PRODUCTS_JSON_URL) as resp:
                    resp.raise_for_status()
                    # content_type may not be application/json -> force parse
                    _PRODUCTS_CACHE = await resp.json(content_type=None)
                    _PRODUCTS_CACHE_TS = time.monotonic()
        return _PRODUCTS_CACHE


def _iter_products(payload: Any):
    """Yield individual product dicts from whatever shape the JSON has."""
    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict):
                yield item
    elif isinstance(payload, dict):
        # Common shapes: {"products": [...]} or {"<id>": {...}, ...}
        if "products" in payload and isinstance(payload["products"], list):
            for item in payload["products"]:
                if isinstance(item, dict):
                    yield item
            return
        for key, item in payload.items():
            if isinstance(item, dict):
                # Inject the key as a candidate id if missing
                item.setdefault("_key", key)
                yield item


_PRODUCT_ID_RE = re.compile(r"product_id=(\d+)")


def _extract_ids(product: dict) -> list[str]:
    """
    Return ALL identifiers a product can be matched on.
    The acoustic.ge feed exposes two distinct numeric IDs:
      - `sku`          (top-level, what staff usually quotes)
      - `product_id`   (extracted from the URL)
    Plus fallbacks for other JSON shapes.
    """
    ids: list[str] = []
    sku = product.get("sku")
    if sku not in (None, ""):
        ids.append(str(sku).strip())

    url = product.get("url") or product.get("URL") or ""
    if isinstance(url, str):
        m = _PRODUCT_ID_RE.search(url)
        if m:
            ids.append(m.group(1))

    for k in ("product_id", "productId", "id", "ID", "Id", "code", "_key"):
        v = product.get(k)
        if v not in (None, ""):
            ids.append(str(v).strip())

    # de-dup while preserving order
    seen: set[str] = set()
    out: list[str] = []
    for i in ids:
        if i not in seen:
            seen.add(i)
            out.append(i)
    return out


def _extract_name(product: dict) -> str:
    for k in ("product", "title", "name", "full_name", "product_name", "Name", "Title"):
        v = product.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _extract_price(product: dict) -> float:
    for k in ("price", "Price", "unit_price", "sale_price", "current_price", "final_price"):
        v = product.get(k)
        if v in (None, ""):
            continue
        try:
            if isinstance(v, str):
                cleaned = re.sub(r"[^\d.,-]", "", v).replace(",", ".")
                return float(cleaned)
            return float(v)
        except (TypeError, ValueError):
            continue
    return 0.0


def _normalize_code(code: str) -> str:
    """Remove all non-alphanumeric characters and normalize case for flexible matching."""
    return re.sub(r'[^A-Za-z0-9]', '', code).upper()


async def fetch_product_data(product_id: str | int) -> dict:
    """
    Return {'id', 'name', 'price'} for the requested product_id.
    Raises ValueError if not found or if multiple variants exist.
    Uses character-agnostic normalization for flexible matching.
    """
    target = str(product_id).strip()
    if not target:
        raise ValueError("Product code cannot be empty")

    # Normalize target (remove all non-alphanumeric characters)
    target_norm = _normalize_code(target)

    payload = await _load_products()

    # Build a list of all products with their normalized IDs
    products_with_ids = []
    for product in _iter_products(payload):
        for pid in _extract_ids(product):
            pid_norm = _normalize_code(pid)
            products_with_ids.append({
                "id": pid,
                "id_norm": pid_norm,
                "name": _extract_name(product),
                "price": _extract_price(product),
            })

    # Pass B: Look for exact match after normalization
    for item in products_with_ids:
        if item["id_norm"] == target_norm:
            return {
                "id": item["id"],
                "name": item["name"],
                "price": item["price"],
            }

    # Pass C: Look for startswith (normalized input)
    candidates = []
    for item in products_with_ids:
        if item["id_norm"].startswith(target_norm):
            candidates.append(item)

    if not candidates:
        raise ValueError(f"Error: Product code {target} not found in database.")

    # Pass D: If multiple matches, ask user to clarify
    if len(candidates) > 1:
        variant_list = ", ".join([f"{c['id']} ({c['name']})" for c in candidates])
        raise ValueError(f"Found multiple variants for this code: {variant_list}. Please specify which one you meant.")

    # Pass E: Only one match, use it automatically
    return {
        "id": candidates[0]["id"],
        "name": candidates[0]["name"],
        "price": candidates[0]["price"],
    }


def _sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\r\n\t]+", " ", name).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "Client"


def _generate_invoice_number() -> str:
    """Generate a unique random invoice number, persisting used numbers to avoid duplicates."""
    used = set()
    if os.path.exists(INVOICE_NUMBERS_FILE):
        try:
            with open(INVOICE_NUMBERS_FILE, "r", encoding="utf-8") as f:
                used = set(json.load(f))
        except (json.JSONDecodeError, IOError):
            pass

    for _ in range(10000):
        num = f"INV-{random.randint(100000, 999999)}"
        if num not in used:
            used.add(num)
            with open(INVOICE_NUMBERS_FILE, "w", encoding="utf-8") as f:
                json.dump(sorted(used), f, indent=2)
            return num

    raise RuntimeError("Could not generate a unique invoice number after 10000 attempts.")


async def generate_invoice(client_info: str, items: list[dict]) -> str:
    """
    Populate the master template with the given client + items, save the
    result into OUTPUT_DIR, and return the absolute path of the new file.

    items: [{'id': '13333', 'qty': 2}, ...]
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    # The template has 4 item rows (8-11). For more items, we dynamically
    # insert extra rows and shift the footer down.
    TEMPLATE_ITEM_ROWS = 4
    extra_rows = max(0, len(items) - TEMPLATE_ITEM_ROWS)

    # Fetch all products concurrently (single HTTP call, just lookups in cache)
    await _load_products()
    resolved = await asyncio.gather(
        *(fetch_product_data(it["id"]) for it in items),
        return_exceptions=True
    )

    # Check for errors (ValueError from fetch_product_data)
    for i, result in enumerate(resolved):
        if isinstance(result, Exception):
            raise result  # Re-raise the ValueError with the specific message

    # Load template directly (no shutil.copy)
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # If more than 4 items, insert extra rows after row 11
    START_ROW = 8
    TEMPLATE_ITEM_ROWS = 4
    LAST_TEMPLATE_ROW = START_ROW + TEMPLATE_ITEM_ROWS - 1  # row 11

    if extra_rows > 0:
        insert_at = LAST_TEMPLATE_ROW + 1  # row 12

        # 1. Unmerge all merged cells at or below the insertion point
        to_remerge = []
        cell_values = {}  # Store cell values before unmerge
        for mr in list(ws.merged_cells.ranges):
            if mr.min_row >= insert_at:
                to_remerge.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
                # Store the value from the top-left cell
                top_left = ws.cell(row=mr.min_row, column=mr.min_col)
                cell_values[(mr.min_row, mr.min_col)] = top_left.value
                ws.unmerge_cells(str(mr))

        # 2. Insert rows
        ws.insert_rows(insert_at, amount=extra_rows)

        # 3. Re-merge with shifted row numbers and preserve row heights
        for min_row, min_col, max_row, max_col in to_remerge:
            new_min_row = min_row + extra_rows
            new_max_row = max_row + extra_rows
            ws.merge_cells(
                start_row=new_min_row, start_column=min_col,
                end_row=new_max_row, end_column=max_col,
            )
            # Restore the cell value
            orig_value = cell_values.get((min_row, min_col))
            if orig_value is not None:
                ws.cell(row=new_min_row, column=min_col, value=orig_value)
            # Preserve row heights for the shifted rows
            for r in range(min_row, max_row + 1):
                orig_height = ws.row_dimensions[r].height
                if orig_height:
                    ws.row_dimensions[r + extra_rows].height = orig_height

        # 4. Copy formatting from row 10 (a full item row) to new rows
        for i in range(extra_rows):
            new_row = insert_at + i
            ws.row_dimensions[new_row].height = ws.row_dimensions[10].height
            for c in range(1, 13):
                src_cell = ws.cell(row=10, column=c)
                dst_cell = ws.cell(row=new_row, column=c)
                if src_cell.has_style:
                    dst_cell.font = src_cell.font.copy()
                    dst_cell.border = src_cell.border.copy()
                    dst_cell.fill = src_cell.fill.copy()
                    dst_cell.number_format = src_cell.number_format
            # Add merged cells for B:H and K:L like other item rows
            ws.merge_cells(start_row=new_row, start_column=2, end_row=new_row, end_column=8)
            ws.merge_cells(start_row=new_row, start_column=11, end_row=new_row, end_column=12)

    # Client name -> A3 (transparently handles merged cells)
    _write_cell(ws, 3, 1, value=client_info)

    # Items table -> starting row 8, clear all item rows before writing
    total_item_rows = TEMPLATE_ITEM_ROWS + extra_rows

    # Clear item table cells (A, B, I, J, K) by setting to empty string
    ITEM_ROW_HEIGHT = 22.75
    for r in range(START_ROW, START_ROW + total_item_rows):
        for c in (1, 2, 9, 10, 11):
            _write_cell(ws, r, c, value="")
        # Ensure all item rows have proper height (row 11 in template is only 7px)
        if ws.row_dimensions[r].height and ws.row_dimensions[r].height < ITEM_ROW_HEIGHT:
            ws.row_dimensions[r].height = ITEM_ROW_HEIGHT

    # Write product data
    for idx, (item, prod) in enumerate(zip(items, resolved), start=1):
        row = START_ROW + idx - 1
        qty = item.get("qty", 0)
        try:
            qty = int(qty)
        except (TypeError, ValueError):
            qty = 0

        price = prod["price"]
        line_total = qty * price

        _write_cell(ws, row, 1, value=idx)              # A: counter
        _write_cell(ws, row, 2, value=prod["name"])     # B: full title
        _write_cell(ws, row, 9, value=qty)              # I: qty
        _write_cell(ws, row, 10, value=price, number_format='#,##0 ₾')   # J: unit price
        _write_cell(ws, row, 11, value=line_total, number_format='#,##0 ₾')  # K: line total

    # Auto-fit column widths for J and K
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15

    # Expand the "მადლობა თანამშრომლობისთვის!" merged cell to span more columns
    # so the text fits without changing column A width
    # Find the row where "მადლობა" is (template row 17, shifted by extra_rows)
    thank_you_row = 17 + extra_rows
    # Unmerge A:E and merge A:K instead
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == thank_you_row and mr.min_col == 1 and mr.max_col == 5:
            ws.unmerge_cells(str(mr))
            ws.merge_cells(start_row=thank_you_row, start_column=1, end_row=thank_you_row, end_column=11)  # A:K
            break
    # Ensure the row is tall enough for the wrapped text to be fully visible
    ws.row_dimensions[thank_you_row].height = 40

    # Also ensure the "ანგარიშის ნომრები" row has enough height
    account_row = 12 + extra_rows
    ws.row_dimensions[account_row].height = 60

    # Calculate Grand Total — row shifts if we inserted extra rows
    grand_total_row = 13 + extra_rows
    grand_total = sum(item.get("qty", 0) * prod["price"] for item, prod in zip(items, resolved))
    _write_cell(ws, grand_total_row, 11, value=grand_total, number_format='#,##0 ₾')

    # Update date — row shifts if we inserted extra rows
    date_row = 18 + extra_rows
    tbilisi_tz = pytz.timezone('Asia/Tbilisi')
    current_date = datetime.now(tbilisi_tz).strftime("%d/%m/%y")
    _write_cell(ws, date_row, 11, value=f"თარიღი: {current_date}")

    # Load images from project folder with exact positioning
    # Remove any existing images to avoid duplicates
    if hasattr(ws, "_images"):
        del ws._images[:]
    
    # Load template to get original image anchors
    template_wb = openpyxl.load_workbook(TEMPLATE_PATH)
    template_ws = template_wb.active

    # Load images from project folder (logo.png and signature.png)
    # with exact positioning matching the template

    # Logo: width 689, height 150, col 5, row 0, offset 59194, 301361
    logo_path = os.path.join(os.path.dirname(TEMPLATE_PATH), "logo.png")
    if os.path.exists(logo_path) and len(template_ws._images) >= 2:
        logo_img = Image(logo_path)
        logo_img.width = 689
        logo_img.height = 150
        # Copy anchor from second image in template (logo)
        logo_img.anchor = template_ws._images[1].anchor
        ws.add_image(logo_img)

    # Signature: width 139, height 46, col 7, row 16, offset 52351, 177086
    signature_path = os.path.join(os.path.dirname(TEMPLATE_PATH), "signature.png")
    if os.path.exists(signature_path) and len(template_ws._images) >= 1:
        signature_img = Image(signature_path)
        signature_img.width = 139
        signature_img.height = 46
        # Position signature at I column, 2 rows above დირექტორი line
        # დირექტორი is at template row 18 (0-indexed 17) + extra_rows
        # User wants it 2 rows above that = 0-indexed (15 + extra_rows)
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
        from copy import deepcopy
        orig_anchor = template_ws._images[0].anchor
        if isinstance(orig_anchor, OneCellAnchor):
            new_anchor = deepcopy(orig_anchor)
            new_anchor._from.row = 15 + extra_rows  # 0-indexed
            new_anchor._from.col = 8  # I column (0-indexed)
            signature_img.anchor = new_anchor
        else:
            signature_img.anchor = orig_anchor
        ws.add_image(signature_img)

    # Generate unique invoice number and write it into the Excel file
    invoice_num = _generate_invoice_number()

    # Write invoice number inside the black banner (A1:F1 merged area)
    # Unmerge A1:F1, keep "ინვოისი" in A1:C1, put invoice number in D1:H1
    # so the number sits closer to "ინვოისი" and is fully visible
    ws.unmerge_cells('A1:F1')
    ws.merge_cells('A1:C1')
    ws.merge_cells('D1:H1')
    a1 = ws.cell(row=1, column=1)
    # Black background with white text for the whole banner
    from openpyxl.styles import PatternFill, Font, Alignment
    black_fill = PatternFill(start_color='FF000000', fill_type='solid')
    for col in range(1, 9):
        cell = ws.cell(row=1, column=col)
        cell.fill = black_fill
        cell.font = Font(name=a1.font.name, size=a1.font.size, bold=a1.font.bold, color='FFFFFF')
    d1 = ws.cell(row=1, column=4)
    d1.value = f"# {invoice_num.replace('INV-', '')}"
    d1.alignment = Alignment(horizontal='left', vertical='center')

    # Save as new file with invoice number in filename
    safe_name = _sanitize_filename(client_info)
    out_path = os.path.abspath(os.path.join(OUTPUT_DIR, f"Invoice_{invoice_num}_{safe_name}.xlsx"))
    wb.save(out_path)
    return out_path
